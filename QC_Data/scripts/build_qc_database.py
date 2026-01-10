#!/usr/bin/env python3
"""
SDP QC Sheets Database Builder
Parses all QC Sheet Excel files and imports them into a SQLite database.
"""

import sqlite3
import os
import glob
from pathlib import Path
from datetime import datetime, date, time, timedelta
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.datetime import from_excel
import re

# Configuration
QC_FORMS_DIR = "/mnt/nvme2/SDP/QC-Data"
DB_PATH = "/mnt/nvme2/SDP/2-Dev/SDP-ProdMgmt2.0/qc_sheets.db"

def create_database_schema(conn):
    """Create the SQLite database schema for QC sheets."""
    cursor = conn.cursor()
    
    # Main QC entries table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS qc_entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            source_file TEXT NOT NULL,
            work_order TEXT,
            customer_name TEXT,
            entry_date DATE,
            operator TEXT,
            part_name TEXT,
            start_time TIME,
            finish_time TIME,
            process_time REAL,
            total_time REAL,
            material TEXT,
            material_size TEXT,
            total_parts INTEGER,
            yield_status TEXT,
            scrap_count INTEGER,
            defects_count INTEGER,
            notes TEXT,
            department TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    
    # File metadata table
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS qc_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL UNIQUE,
            file_path TEXT NOT NULL,
            file_size INTEGER,
            last_modified TIMESTAMP,
            total_entries INTEGER DEFAULT 0,
            imported_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            import_status TEXT DEFAULT 'success',
            error_message TEXT
        )
    """)
    
    # Create indexes for common queries
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_entry_date ON qc_entries(entry_date)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_operator ON qc_entries(operator)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_part_name ON qc_entries(part_name)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_work_order ON qc_entries(work_order)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_department ON qc_entries(department)")
    cursor.execute("CREATE INDEX IF NOT EXISTS idx_source_file ON qc_entries(source_file)")
    
    conn.commit()
    print("Database schema created successfully.")

def parse_filename(filename):
    """Extract work order and customer name from filename."""
    # Remove extension
    name = Path(filename).stem
    
    # Common patterns: "CUSTOMER PO 12345", "CUSTOMER 12345", etc.
    # Try to extract PO/work order number
    po_match = re.search(r'(PO|PO\s*#?)\s*([A-Z0-9\-]+)', name, re.IGNORECASE)
    work_order = po_match.group(2) if po_match else None
    
    # Extract customer name (everything before PO or number)
    if po_match:
        customer = name[:po_match.start()].strip()
    else:
        # Try to find customer name (usually first part before numbers)
        parts = re.split(r'\s+(\d+)', name, 1)
        customer = parts[0].strip() if parts else name
    
    return {
        'work_order': work_order,
        'customer_name': customer if customer else None
    }

def detect_department(operator, part_name, material):
    """Try to detect department from operator code or other indicators."""
    if not operator:
        return None
    
    operator_upper = operator.upper()
    
    # Based on Employees.md and common patterns
    if 'ASSEMBLY' in operator_upper or 'ASSY' in str(part_name).upper():
        return 'Assembly'
    elif 'WATERJET' in operator_upper or 'WJ' in operator_upper:
        return 'WaterJet'
    elif 'ROUTER' in operator_upper or 'RT' in operator_upper:
        return 'Router'
    elif 'SAW' in operator_upper:
        return 'Saw'
    elif 'PRESS' in operator_upper:
        return 'Press'
    
    return None

def parse_yield_status(yield_value):
    """Parse yield status and extract scrap/defect counts."""
    if not yield_value:
        return None, None, None
    
    yield_str = str(yield_value).upper()
    
    # Check for scrap
    if 'SCRAP' in yield_str:
        # Try to extract number
        scrap_match = re.search(r'(\d+)', yield_str)
        scrap_count = int(scrap_match.group(1)) if scrap_match else 1
        return 'SCRAP', scrap_count, None
    
    # Check for defects
    if 'DEFECT' in yield_str or 'BAD' in yield_str:
        defect_match = re.search(r'(\d+)', yield_str)
        defect_count = int(defect_match.group(1)) if defect_match else 1
        return 'DEFECT', None, defect_count
    
    return yield_str, None, None

def parse_excel_file(file_path, conn):
    """Parse a single Excel file and import its data."""
    cursor = conn.cursor()
    filename = os.path.basename(file_path)
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb.active
        
        # Get file metadata
        file_stat = os.stat(file_path)
        file_info = parse_filename(filename)
        
        # Find header row
        header_row_idx = None
        headers = {}
        
        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if not any(cell for cell in row):
                continue
            
            # Check if this looks like a header row
            row_str = ' '.join(str(cell).upper() if cell else '' for cell in row[:15])
            if 'DATE' in row_str and 'OPERATOR' in row_str:
                header_row_idx = row_idx
                # Map headers
                for col_idx, header in enumerate(row):
                    if header:
                        header_clean = str(header).strip().upper()
                        headers[header_clean] = col_idx
                break
        
        if not header_row_idx:
            raise ValueError("Could not find header row")
        
        # Parse data rows (skip header row)
        entries = []
        all_rows = list(ws.iter_rows(values_only=True))
        # header_row_idx from enumerate is 1-indexed, but when used as list index it works correctly
        # because enumerate(iter, 1) numbers items starting at 1, matching their 0-indexed position + 1
        # So header_row_idx=6 means the 6th item (at list index 5), and we want to start from index 6
        for i, row in enumerate(all_rows[header_row_idx:], start=header_row_idx + 1):
            row_idx = i  # 1-indexed row number for error messages
            # Skip empty rows
            if not any(cell for cell in row):
                continue
            
            # Check if this is a data row (has date and operator)
            # Find DATE column (header might be "DATE (mm/dd/yyyy)" or similar)
            date_col_idx = None
            for key in headers.keys():
                if 'DATE' in key:
                    date_col_idx = headers[key]
                    break
            if date_col_idx is None:
                date_col_idx = 0  # Default to first column
            
            # Find OPERATOR column
            operator_col_idx = None
            for key in headers.keys():
                if 'OPERATOR' in key:
                    operator_col_idx = headers[key]
                    break
            if operator_col_idx is None:
                operator_col_idx = 1  # Default to second column
            
            if date_col_idx >= len(row):
                continue
            
            date_val = row[date_col_idx]
            operator_val = row[operator_col_idx] if operator_col_idx < len(row) else None
            
            # Require date, but operator can be empty (will be set to None)
            # Check if date_val is None, empty string, or falsy
            if date_val is None or (isinstance(date_val, str) and not date_val.strip()):
                continue
            
            # Parse row data
            try:
                entry_date = None
                if isinstance(date_val, datetime):
                    entry_date = date_val.date()
                elif isinstance(date_val, date):
                    entry_date = date_val
                elif isinstance(date_val, (int, float)):
                    # Handle Excel serial number dates
                    # When openpyxl reads dates with data_only=True, if cells aren't formatted
                    # as dates in Excel, they're returned as floats (Excel serial numbers)
                    try:
                        # Use openpyxl's built-in conversion which handles Excel's date system correctly
                        # (accounts for Excel incorrectly treating 1900 as a leap year)
                        excel_datetime = from_excel(date_val)
                        if isinstance(excel_datetime, datetime):
                            entry_date = excel_datetime.date()
                        elif isinstance(excel_datetime, date):
                            entry_date = excel_datetime
                        else:
                            entry_date = None
                        
                        # Validate the date is reasonable (between 2000 and 2100 for QC sheets)
                        # This filters out obviously wrong dates like 1900 or 2925
                        if entry_date and (entry_date.year < 2000 or entry_date.year > 2100):
                            # Try alternative: maybe it's a Unix timestamp (seconds since 1970)
                            if date_val > 1000000000:  # Likely Unix timestamp
                                try:
                                    entry_date = datetime.fromtimestamp(date_val).date()
                                    # Re-validate after timestamp conversion
                                    if entry_date.year < 2000 or entry_date.year > 2100:
                                        entry_date = None
                                except:
                                    entry_date = None
                            else:
                                entry_date = None
                    except Exception as e:
                        entry_date = None
                elif isinstance(date_val, str) and date_val.strip():
                    # Try to parse date string
                    # Handle dates with time components by splitting on space
                    date_str = date_val.strip()
                    if ' ' in date_str:
                        date_str = date_str.split()[0]  # Take just the date part
                    
                    # Try multiple date formats (including formats with time)
                    date_formats = [
                        '%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%Y/%m/%d', 
                        '%m-%d-%Y', '%d-%m-%Y', '%Y-%m-%d %H:%M:%S',
                        '%m/%d/%Y %H:%M:%S', '%d/%m/%Y %H:%M:%S'
                    ]
                    for fmt in date_formats:
                        try:
                            entry_date = datetime.strptime(date_str, fmt).date()
                            break
                        except:
                            continue
                
                # If date parsing failed, skip this row
                if not entry_date:
                    continue
                
                operator = str(operator_val).strip() if operator_val else None
                part_name = str(row[headers.get('PART NAME', 2)]) if 'PART NAME' in headers and row[headers.get('PART NAME', 2)] else None
                
                start_time = None
                start_val = row[headers.get('START', 3)] if 'START' in headers else None
                if isinstance(start_val, time):
                    start_time = start_val
                elif isinstance(start_val, datetime):
                    start_time = start_val.time()
                
                finish_time = None
                finish_val = row[headers.get('FINISH', 5)] if 'FINISH' in headers else None
                if isinstance(finish_val, time):
                    finish_time = finish_val
                elif isinstance(finish_val, datetime):
                    finish_time = finish_val.time()
                
                process_time = row[headers.get('PROCESS TIME', 4)] if 'PROCESS TIME' in headers else None
                total_time = row[headers.get('TOTAL TIME', 6)] if 'TOTAL TIME' in headers else None
                
                material = str(row[headers.get('MATERIAL', 7)]) if 'MATERIAL' in headers and row[headers.get('MATERIAL', 7)] else None
                material_size = str(row[headers.get('MATERIAL SIZE', 10)]) if 'MATERIAL SIZE' in headers and row[headers.get('MATERIAL SIZE', 10)] else None
                
                total_parts = None
                parts_val = row[headers.get('TOTAL PARTS', 8)] if 'TOTAL PARTS' in headers else None
                if isinstance(parts_val, (int, float)):
                    total_parts = int(parts_val)
                elif parts_val:
                    try:
                        total_parts = int(float(str(parts_val)))
                    except:
                        pass
                
                yield_val = row[headers.get('YIELD', 9)] if 'YIELD' in headers else None
                yield_status, scrap_count, defects_count = parse_yield_status(yield_val)
                
                # Detect department
                department = detect_department(operator, part_name, material)
                
                # Convert date/time objects to strings for SQLite
                entry_date_str = entry_date.isoformat() if entry_date else None
                start_time_str = start_time.strftime('%H:%M:%S') if start_time else None
                finish_time_str = finish_time.strftime('%H:%M:%S') if finish_time else None
                
                entry = {
                    'source_file': filename,
                    'work_order': file_info['work_order'],
                    'customer_name': file_info['customer_name'],
                    'entry_date': entry_date_str,
                    'operator': operator,
                    'part_name': part_name,
                    'start_time': start_time_str,
                    'finish_time': finish_time_str,
                    'process_time': float(process_time) if process_time and isinstance(process_time, (int, float)) else None,
                    'total_time': float(total_time) if total_time and isinstance(total_time, (int, float)) else None,
                    'material': material,
                    'material_size': material_size,
                    'total_parts': total_parts,
                    'yield_status': yield_status,
                    'scrap_count': scrap_count,
                    'defects_count': defects_count,
                    'department': department
                }
                
                entries.append(entry)
                
            except Exception as e:
                print(f"  Warning: Error parsing row {row_idx + 1}: {e}")
                continue
        
        # Delete existing entries for this file to prevent duplicates on re-import
        cursor.execute("DELETE FROM qc_entries WHERE source_file = ?", (filename,))
        
        # Insert entries into database
        insert_count = 0
        for entry in entries:
            try:
                cursor.execute("""
                    INSERT INTO qc_entries (
                        source_file, work_order, customer_name, entry_date, operator,
                        part_name, start_time, finish_time, process_time, total_time,
                        material, material_size, total_parts, yield_status, scrap_count,
                        defects_count, department
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """, (
                    entry['source_file'],
                    entry['work_order'],
                    entry['customer_name'],
                    entry['entry_date'],
                    entry['operator'],
                    entry['part_name'],
                    entry['start_time'],
                    entry['finish_time'],
                    entry['process_time'],
                    entry['total_time'],
                    entry['material'],
                    entry['material_size'],
                    entry['total_parts'],
                    entry['yield_status'],
                    entry['scrap_count'],
                    entry['defects_count'],
                    entry['department']
                ))
                insert_count += 1
            except Exception as e:
                print(f"  Warning: Error inserting entry: {e}")
                continue
        
        # Record file import
        cursor.execute("""
            INSERT OR REPLACE INTO qc_files (
                filename, file_path, file_size, last_modified, total_entries,
                import_status, error_message
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
        """, (
            filename,
            file_path,
            file_stat.st_size,
            datetime.fromtimestamp(file_stat.st_mtime),
            insert_count,
            'success',
            None
        ))
        
        conn.commit()
        return insert_count, None
        
    except Exception as e:
        error_msg = str(e)
        # Record failed import
        try:
            file_stat = os.stat(file_path)
            cursor.execute("""
                INSERT OR REPLACE INTO qc_files (
                    filename, file_path, file_size, last_modified,
                    import_status, error_message
                ) VALUES (?, ?, ?, ?, ?, ?)
            """, (
                filename,
                file_path,
                file_stat.st_size,
                datetime.fromtimestamp(file_stat.st_mtime),
                'error',
                error_msg
            ))
            conn.commit()
        except:
            pass
        
        return 0, error_msg

def main():
    """Main function to build the QC database."""
    print("=" * 60)
    print("SDP QC Sheets Database Builder")
    print("=" * 60)
    
    # Check if QC Forms directory exists
    if not os.path.exists(QC_FORMS_DIR):
        print(f"Error: QC Forms directory not found: {QC_FORMS_DIR}")
        return
    
    # Find all Excel files
    excel_files = glob.glob(os.path.join(QC_FORMS_DIR, "*.xlsx"))
    excel_files.extend(glob.glob(os.path.join(QC_FORMS_DIR, "*.xls")))
    
    print(f"\nFound {len(excel_files)} Excel files to process")
    
    # Create/connect to database
    conn = sqlite3.connect(DB_PATH)
    create_database_schema(conn)
    
    # Process files
    total_entries = 0
    successful_files = 0
    failed_files = 0
    
    print("\nProcessing files...")
    for i, file_path in enumerate(excel_files, 1):
        filename = os.path.basename(file_path)
        print(f"[{i}/{len(excel_files)}] Processing: {filename}")
        
        count, error = parse_excel_file(file_path, conn)
        
        if error:
            print(f"  ❌ Error: {error}")
            failed_files += 1
        else:
            print(f"  ✅ Imported {count} entries")
            total_entries += count
            successful_files += 1
    
    # Print summary
    print("\n" + "=" * 60)
    print("Import Summary")
    print("=" * 60)
    print(f"Total files processed: {len(excel_files)}")
    print(f"Successful: {successful_files}")
    print(f"Failed: {failed_files}")
    print(f"Total QC entries imported: {total_entries}")
    print(f"\nDatabase saved to: {DB_PATH}")
    
    # Print some statistics
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(DISTINCT operator) FROM qc_entries WHERE operator IS NOT NULL")
    operator_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT COUNT(DISTINCT work_order) FROM qc_entries WHERE work_order IS NOT NULL")
    wo_count = cursor.fetchone()[0]
    
    cursor.execute("SELECT MIN(entry_date), MAX(entry_date) FROM qc_entries WHERE entry_date IS NOT NULL")
    date_range = cursor.fetchone()
    
    print(f"\nDatabase Statistics:")
    print(f"  Unique operators: {operator_count}")
    print(f"  Unique work orders: {wo_count}")
    if date_range[0] and date_range[1]:
        print(f"  Date range: {date_range[0]} to {date_range[1]}")
    
    conn.close()
    print("\n✅ Database build complete!")

if __name__ == "__main__":
    main()
